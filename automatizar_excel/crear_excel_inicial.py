import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import openpyxl.utils


def crear_excel_inicial():
    # Crear un nuevo libro de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario Principal"

    # Definir encabezados
    encabezados = ["ID", "Producto", "Categoria", "Precio", "Stock"]

    # Datos random
    datos = [
        [1, "Laptop", "Libros", 98.19, 61],
        [2, "Cafetera", "Electrónica", 292.59, 71],
        [3, "Impresora", "Hogar", 881.11, 9],
        [4, "Teléfono", "Ropa", 1330.07, 88],
        [5, "Libro", "Ropa", 341.72, 1],
        [6, "Impresora", "Accesorios", 961.78, 62],
        [7, "Teclado", "Ropa", 376.54, 23],
        [8, "Auriculares", "Hogar", 264.6, 43],
        [9, "Teclado", "Electrónica", 1018.06, 77],
        [10, "Cafetera", "Electrónica", 1002.96, 82],
        [11, "Ratón", "Libros", 1453.15, 10],
        [12, "Monitor", "Electrónica", 172.49, 4],
        [13, "Teclado", "Ropa", 400.15, 95],
        [14, "Cafetera", "Hogar", 830.14, 98],
        [15, "Camiseta", "Accesorios", 953.26, 92],
        [16, "Monitor", "Hogar", 1031.8, 32],
        [17, "Cafetera", "Libros", 886.57, 38],
        [18, "Teclado", "Accesorios", 1183.42, 24],
        [19, "Teléfono", "Accesorios", 602.53, 53],
        [20, "Camiseta", "Electrónica", 526.68, 84],
        [21, "Laptop", "Electrónica", 1200.00, 15],
        [22, "Camiseta", "Ropa", 25.00, 50],
        [23, "Cafetera", "Hogar", 80.00, 30],
        [24, "Libro", "Libros", 15.00, 100],
        [25, "Teléfono", "Electrónica", 800.00, 20],
        [26, "Auriculares", "Accesorios", 50.00, 40],
        [27, "Monitor", "Electrónica", 300.00, 25],
        [28, "Teclado", "Accesorios", 45.00, 60],
        [29, "Ratón", "Electrónica", 20.00, 70],
        [30, "Impresora", "Hogar", 150.00, 10],
        [31, "Laptop", "Electrónica", 1100.00, 12],
        [32, "Camiseta", "Ropa", 30.00, 55],
        [33, "Cafetera", "Hogar", 90.00, 35],
        [34, "Libro", "Libros", 20.00, 90],
        [35, "Teléfono", "Electrónica", 850.00, 18],
        [36, "Auriculares", "Accesorios", 60.00, 45],
        [37, "Monitor", "Electrónica", 320.00, 22],
        [38, "Teclado", "Accesorios", 50.00, 65],
        [39, "Ratón", "Electrónica", 25.00, 75],
        [40, "Impresora", "Hogar", 160.00, 8],
        [41, "Laptop", "Electrónica", 1150.00, 14],
        [42, "Camiseta", "Ropa", 35.00, 60],
        [43, "Cafetera", "Hogar", 100.00, 40],
        [44, "Libro", "Libros", 25.00, 85],
        [45, "Teléfono", "Electrónica", 900.00, 16],
        [46, "Auriculares", "Accesorios", 70.00, 50],
        [47, "Monitor", "Electrónica", 340.00, 20],
        [48, "Teclado", "Accesorios", 55.00, 70],
        [49, "Ratón", "Electrónica", 30.00, 80],
        [50, "Impresora", "Hogar", 170.00, 6],
        [51, "Laptop", "Electrónica", 1250.00, 10],
        [52, "Camiseta", "Ropa", 40.00, 65],
        [53, "Cafetera", "Hogar", 110.00, 45],
        [54, "Libro", "Libros", 30.00, 80],
        [55, "Teléfono", "Electrónica", 950.00, 14],
        [56, "Auriculares", "Accesorios", 80.00, 55],
        [57, "Monitor", "Electrónica", 360.00, 18],
        [58, "Teclado", "Accesorios", 60.00, 75],
        [59, "Ratón", "Electrónica", 35.00, 85],
        [60, "Impresora", "Hogar", 180.00, 4],
        [61, "Laptop", "Electrónica", 1300.00, 8],
        [62, "Camiseta", "Ropa", 45.00, 70],
        [63, "Cafetera", "Hogar", 120.00, 50],
        [64, "Libro", "Libros", 35.00, 75],
        [65, "Teléfono", "Electrónica", 1000.00, 12],
        [66, "Auriculares", "Accesorios", 90.00, 60],
        [67, "Monitor", "Electrónica", 380.00, 16],
        [68, "Teclado", "Accesorios", 65.00, 80],
        [69, "Ratón", "Electrónica", 40.00, 90],
        [70, "Impresora", "Hogar", 190.00, 2],
        [71, "Laptop", "Electrónica", 1350.00, 6],
        [72, "Camiseta", "Ropa", 50.00, 75],
        [73, "Cafetera", "Hogar", 130.00, 55],
        [74, "Libro", "Libros", 40.00, 70],
        [75, "Teléfono", "Electrónica", 1050.00, 10],
        [76, "Auriculares", "Accesorios", 100.00, 65],
        [77, "Monitor", "Electrónica", 400.00, 14],
        [78, "Teclado", "Accesorios", 70.00, 85],
        [79, "Ratón", "Electrónica", 45.00, 95],
        [80, "Impresora", "Hogar", 200.00, 1],
        [81, "Laptop", "Electrónica", 1400.00, 4],
        [82, "Camiseta", "Ropa", 55.00, 80],
        [83, "Cafetera", "Hogar", 140.00, 60],
        [84, "Libro", "Libros", 45.00, 65],
        [85, "Teléfono", "Electrónica", 1100.00, 8],
        [86, "Auriculares", "Accesorios", 110.00, 70],
        [87, "Monitor", "Electrónica", 420.00, 12],
        [88, "Teclado", "Accesorios", 75.00, 90],
        [89, "Ratón", "Electrónica", 50.00, 100],
        [90, "Impresora", "Hogar", 210.00, 3],
        [91, "Laptop", "Electrónica", 1450.00, 2],
        [92, "Camiseta", "Ropa", 60.00, 85],
        [93, "Cafetera", "Hogar", 150.00, 65],
        [94, "Libro", "Libros", 50.00, 60],
        [95, "Teléfono", "Electrónica", 1150.00, 6],
        [96, "Auriculares", "Accesorios", 120.00,2]
    ]

    # Escribir encabezados
    for col, encabezado in enumerate(encabezados, 1):
        #print(f"Col = {col}\nEncabezado= {encabezado}")
        celda = ws.cell(row=1, column=col)
        celda.value = encabezado

        # Formato / Estilos
        celda.font = Font(bold=True)
        celda.fill = PatternFill(start_color="CCCCCC", fill_type="solid") # darkGrid para hacer una cuadricula
        celda.alignment = Alignment(horizontal="center")

    # Escribir datos
    for row_idx, fila in enumerate(datos,2):
        for col_idx, valor in enumerate(fila,1):
            celda = ws.cell(row=row_idx, column=col_idx)
            celda.value = valor
            celda.alignment = Alignment(horizontal="center")


    # Ajustar ancho de columna
    for col in range(1, len(encabezados)+1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25







    # Guardar
    wb.save('inventario_tecnologia.xlsx')

    print("Archivo crear exitosamente")


if __name__ == '__main__':
    crear_excel_inicial()