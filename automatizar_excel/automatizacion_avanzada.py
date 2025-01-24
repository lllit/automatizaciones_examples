import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

import os

def crear_graficos_ventas(ws):
    # Crear graficos para stock

    # Crear grafico barras
    grafico = BarChart()
    grafico.y_axis.title = 'Cantidad en stock'
    grafico.x_axis.title = 'Productos'

    # Datos para el grafico
    row_stats = ws.max_row
    col_stats = ws.max_column
    
    datos = Reference(ws, min_col=5, min_row=1, max_row=ws.max_row, max_col=5)
    categorias = Reference(ws, min_col=2,min_row=2,max_row=ws.max_row)

    grafico.add_data(datos,titles_from_data=True)
    grafico.set_categories(categorias)

    # Añador graf a la hoja
    ws.add_chart(grafico, "G2")

def crear_grafico_categorias(ws):
    # crear grafico circular
    grafico = PieChart()
    grafico.title = "Distribución por Categorias"

    datos = Reference(ws, min_col=5, min_row=1, max_row=ws.max_row, max_col=5)
    etiquetas = Reference(ws, min_col=3,min_row=2,max_row=ws.max_row)

    grafico.add_data(datos,titles_from_data=True)
    grafico.set_categories(etiquetas)

    ws.add_chart(grafico,"G18")


def crear_tabla(wb, ws):
    # Crear tabla
    # covnertir rango datos en tabla
    tab = Table(displayName="TablaInventario", ref=f"A1:E{ws.max_row}")

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)


def aplicar_validacion_datos(ws):
    # Aplicar validacion

    # Categorias
    categorias_val = DataValidation(
        type="list",
        formula1="'Computadoras, Smartphones, Accesorios, Monitoreos, Periféricos'",
        allow_blank=False
    )
    categorias_val.error = "Por favor seleccione una categoria valida"
    categorias_val.errorTitle = "Categoria Invalida"

    ws.add_data_validation(categorias_val)
    categorias_val.add(f"C2:C{ws.max_row}")


def aplicar_validacion_datos_stock(ws):
    stock_val = DataValidation(
        type="whole",
        operator="greaterThan",
        formula1="0"
    )
    stock_val.error = "El stock debe ser numero positivo"
    stock_val.errorTitle = "Stock Invalido"

    ws.add_data_validation(stock_val)
    stock_val.add(f"E2:E{ws.max_row}")


def aplicar_formatos_condicionales(ws):
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

    # Escala de colores para precios
    color_scale = ColorScaleRule(
        start_type='min', start_color="90EE90",
        end_type='max', end_color="FF6B6B"
    )
    ws.conditional_formatting.add(f"D2:D{ws.max_row}", color_scale)
    

    

def automatizacion_avanzada():
    # Funcion principal 
    #Cargar archivo exitesnte
    wb = openpyxl.load_workbook("inventario_tecnologia.xlsx")
    ws = wb.active

    print("\n=== Inciando Automatizacion Avanzada ===")

    # Aplicar todas las funcionalidades
    print("1. Creando graficos automaticos...")
    crear_graficos_ventas(ws)
    crear_grafico_categorias(ws)


    print("2. Creando tabla...")
    crear_tabla(wb,ws)

    #print("3. Aplicando validacion de datos...")
    #aplicar_validacion_datos(ws)

    print("3. Aplicando validacion de datos stock...")
    aplicar_validacion_datos_stock(ws)


    print("4. Aplicando formatos condicionales...")
    aplicar_formatos_condicionales(ws)

    #Guardar cambios
    wb.save('inventario_tecnologia.xlsx')

if __name__ == "__main__":
    automatizacion_avanzada()




