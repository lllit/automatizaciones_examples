import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import datetime

def cargar_inventario():
    try:
        wb = openpyxl.load_workbook('inventario_tecnologia.xlsx')

    except FileNotFoundError:
        print("#Error")
        return None
    
    ws = wb.active
    ws.title = "Inventario Principal"

    # Ajustar ancho de columna
    for col in range(1, ws.max_column+1):
        celda = ws.cell(row=1, column=col)
        celda.font = Font(bold=True)
        celda.fill = PatternFill(start_color="CCCCCC", fill_type="solid")
        celda.alignment = Alignment(horizontal="center")

    return wb, ws

def actualizar_precios(ws, porcentaje):
    # Actualiza los precios segun porcentaje
    for row in range(2, ws.max_row+1):
        precio_actual = float(ws.cell(row=row, column=4).value)
        nuevo_precio = precio_actual * (1 + porcentaje /100)
        ws.cell(row=row, column=4).value = round(nuevo_precio, 2)


def verificar_stock_bajo(ws, limite_minimo=10):
    # Generar alertas para productos con bajo stock
    productos_bajos = []
    for row in range(2,ws.max_row +1):
        stock = int(ws.cell(row= row, column=5).value)
        if stock <= limite_minimo:
            producto = ws.cell(row=row, column=2).value
            productos_bajos.append((producto,stock))
            
            # Marcar en rojo las filas que tienen stock bajo el limite
            for col in range(1, ws.max_column +1):
                ws.cell(
                    row=row,
                    column=col
                ).fill = PatternFill(
                    start_color="FFB6B6",
                    fill_type="solid"
                )
    return productos_bajos        

def generar_reporte():
    # Genera reporte inventario actual
    wb, ws = cargar_inventario()
    if not wb:
        return
    
    # Crear nueva hoja
    fecha_actual = datetime.datetime.now().strftime("%Y-%m-%d")
    ws_reporte = wb.create_sheet(f"Reporte {fecha_actual}")

    #Copiar encabezados
    for col in range(1, ws.max_column + 1):
        ws_reporte.cell(row=1, column=col).value = ws.cell(row=1, column=col).value
        ws_reporte.cell(row=1, column=col).font = Font(bold=True)

    # Copiar datos
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws_reporte.cell(row=row, column=col).value = ws.cell(row=row, column=col).value

    # Añadir estadisticas
    row_stats = ws.max_row + 2
    col_stats = ws.max_column + 2

    ws_reporte.cell(row=row_stats, column=1, value= "Estadisticas del inventario").font = Font(bold=True)
    #ws_reporte.cell(row=1, column=col_stats, value= "Estadisticas del inventario").font = Font(bold=True)


    # Productos
    ws_reporte.cell(row=row_stats+1, column=1, value= "Productos: ")
    ws_reporte.cell(row=row_stats+1, column=2, value=f"=COUNTA(B2:B{ws.max_row})")

    # Total inventario
    ws_reporte.cell(row=row_stats+2, column=1, value="Total del inventario (stock): ")
    ws_reporte.cell(row=row_stats+2, column=2, value=f"=SUM(E2:E{ws.max_row})")

    # Valor total inventario
    ws_reporte.cell(row=row_stats+3, column=1, value="Valor total inventario")
    ws_reporte.cell(row=row_stats+3, column=2, value=f"=SUMPRODUCT(D2:D{ws.max_row}, E2:E{ws.max_row})")


    return wb

def automatizacion_inventario():
    # funcion princiapl ejecuta operacion de automatizacion
    wb, ws = cargar_inventario()
    
    if not wb:
        return
    print("\n=== Sistema de Automatización de Inventario ===")

    #Actualizar precios
    incremento = 5
    actualizar_precios(ws,incremento)
    print(f'\nPrecios Actualizados con un incremento del {incremento}%')
    # -------------------------

    # Verificar stock bajo
    productos_bajos = verificar_stock_bajo(ws)
    if productos_bajos:
        print("\n¡Alerta! Productos con stock bajo:")
        for producto, stock in productos_bajos:
            print(f"- {producto}: {stock} unidades")

    # ---------------------

    wb = generar_reporte()

    # --------------------
    wb.save('inventario_tecnologia.xlsx')
    print("\nProceso Completado")


if __name__ == '__main__':
    automatizacion_inventario()